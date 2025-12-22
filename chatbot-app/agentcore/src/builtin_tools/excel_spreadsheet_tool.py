"""
Excel Spreadsheet Tools - 4 essential tools for Excel spreadsheet management.

Tools:
1. create_excel_spreadsheet - Create new Excel spreadsheet from Python code
2. modify_excel_spreadsheet - Modify existing Excel spreadsheet with openpyxl code
3. list_my_excel_spreadsheets - List all Excel spreadsheets in workspace
4. read_excel_spreadsheet - Retrieve spreadsheet for download

Note: Uploaded .xlsx files are automatically stored to workspace by agent.py
Pattern follows word_document_tool for Code Interpreter usage.
"""

import os
import re
import logging
from typing import Dict, Any, Optional
from strands import tool, ToolContext
from bedrock_agentcore.tools.code_interpreter_client import CodeInterpreter
from .lib.document_manager import ExcelDocumentManager

logger = logging.getLogger(__name__)


def _validate_spreadsheet_name(name: str) -> tuple[bool, Optional[str]]:
    """Validate spreadsheet name meets requirements (without extension).

    Rules:
    - Only letters (a-z, A-Z), numbers (0-9), and hyphens (-)
    - No spaces, underscores, or special characters
    - No consecutive hyphens
    - No leading/trailing hyphens

    Args:
        name: Spreadsheet name without extension (e.g., "sales-report")

    Returns:
        (is_valid, error_message)
        - (True, None) if valid
        - (False, error_message) if invalid
    """
    # Check for empty name
    if not name:
        return False, "Spreadsheet name cannot be empty"

    # Check for valid characters: only letters, numbers, hyphens
    if not re.match(r'^[a-zA-Z0-9\-]+$', name):
        invalid_chars = re.findall(r'[^a-zA-Z0-9\-]', name)
        return False, f"Invalid characters in name: {set(invalid_chars)}. Use only letters, numbers, and hyphens (-)."

    # Check for consecutive hyphens
    if '--' in name:
        return False, "Name cannot contain consecutive hyphens (--)"

    # Check for leading/trailing hyphens
    if name.startswith('-') or name.endswith('-'):
        return False, "Name cannot start or end with a hyphen"

    return True, None


def _sanitize_spreadsheet_name_for_bedrock(filename: str) -> str:
    """Sanitize existing filename for Bedrock API (removes extension).

    Use this ONLY for existing files being read from S3.
    For new files, use _validate_spreadsheet_name() instead.

    Args:
        filename: Original filename with extension (e.g., "test_spreadsheet_v2.xlsx")

    Returns:
        Sanitized name without extension (e.g., "test-spreadsheet-v2")
    """
    # Remove extension
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
    else:
        name = filename

    # Replace underscores and spaces with hyphens
    name = name.replace('_', '-').replace(' ', '-')

    # Keep only alphanumeric and hyphens
    name = re.sub(r'[^a-zA-Z0-9\-]', '', name)

    # Replace multiple consecutive hyphens with single hyphen
    name = re.sub(r'\-+', '-', name)

    # Trim hyphens from start/end
    name = name.strip('-')

    # If name becomes empty, use default
    if not name:
        name = 'spreadsheet'

    if name != filename.replace('.xlsx', ''):
        logger.info(f"Sanitized spreadsheet name for Bedrock: '{filename}' → '{name}'")

    return name


def _get_code_interpreter_id() -> Optional[str]:
    """Get Custom Code Interpreter ID from environment or Parameter Store"""
    # 1. Check environment variable (set by AgentCore Runtime)
    code_interpreter_id = os.getenv('CODE_INTERPRETER_ID')
    if code_interpreter_id:
        logger.info(f"Found CODE_INTERPRETER_ID in environment: {code_interpreter_id}")
        return code_interpreter_id

    # 2. Try Parameter Store (for local development or alternative configuration)
    try:
        import boto3
        project_name = os.getenv('PROJECT_NAME', 'strands-agent-chatbot')
        environment = os.getenv('ENVIRONMENT', 'dev')
        region = os.getenv('AWS_REGION', 'us-west-2')
        param_name = f"/{project_name}/{environment}/agentcore/code-interpreter-id"

        logger.info(f"Checking Parameter Store for Code Interpreter ID: {param_name}")
        ssm = boto3.client('ssm', region_name=region)
        response = ssm.get_parameter(Name=param_name)
        code_interpreter_id = response['Parameter']['Value']
        logger.info(f"Found CODE_INTERPRETER_ID in Parameter Store: {code_interpreter_id}")
        return code_interpreter_id
    except Exception as e:
        logger.warning(f"Custom Code Interpreter ID not found in Parameter Store: {e}")
        return None


def _get_user_session_ids(tool_context: ToolContext) -> tuple[str, str]:
    """Extract user_id and session_id from ToolContext

    Returns:
        (user_id, session_id) tuple
    """
    # Extract from invocation_state (set by agent)
    invocation_state = tool_context.invocation_state
    user_id = invocation_state.get('user_id', 'default_user')
    session_id = invocation_state.get('session_id', 'default_session')

    logger.info(f"Extracted IDs: user_id={user_id}, session_id={session_id}")
    return user_id, session_id


def _load_workspace_images_to_ci(code_interpreter: CodeInterpreter, user_id: str, session_id: str) -> list[str]:
    """Load all images from S3 workspace to Code Interpreter

    This replaces the old invocation_state approach with S3-based storage.
    Images are now stored in S3 workspace and loaded on-demand when tools execute.

    Args:
        code_interpreter: Active CodeInterpreter instance
        user_id: User ID for S3 workspace path
        session_id: Session ID for S3 workspace path

    Returns:
        List of loaded image filenames
    """
    from .lib.document_manager import ImageDocumentManager

    try:
        # Initialize image document manager
        image_manager = ImageDocumentManager(user_id, session_id)

        # List all images from S3 workspace
        images = image_manager.list_s3_documents()

        if not images:
            logger.info("No images found in workspace")
            return []

        logger.info(f"Found {len(images)} image(s) in workspace: {[img['filename'] for img in images]}")

        # Load each image from S3 and upload to Code Interpreter
        loaded_filenames = []
        for image_info in images:
            try:
                filename = image_info['filename']

                # Load image bytes from S3
                file_bytes = image_manager.load_from_s3(filename)

                # Upload to Code Interpreter using existing helper
                _upload_images_to_ci(code_interpreter, [(filename, file_bytes)])

                loaded_filenames.append(filename)
                logger.info(f"✅ Loaded image from S3 workspace: {filename}")

            except Exception as e:
                logger.error(f"Failed to load image {filename} from workspace: {e}")
                continue

        return loaded_filenames

    except Exception as e:
        logger.error(f"Failed to load workspace images: {e}")
        return []


def _upload_images_to_ci(code_interpreter: CodeInterpreter, images: list[tuple[str, bytes]]) -> list[str]:
    """Upload multiple image files to Code Interpreter workspace

    Args:
        code_interpreter: Active CodeInterpreter instance
        images: List of (filename, file_bytes) tuples

    Returns:
        List of uploaded filenames
    """
    uploaded = []

    for filename, file_bytes in images:
        try:
            import base64
            encoded_bytes = base64.b64encode(file_bytes).decode('utf-8')

            write_code = f"""
import base64

# Decode and write image file
file_bytes = base64.b64decode('{encoded_bytes}')
with open('{filename}', 'wb') as f:
    f.write(file_bytes)

print(f"Image uploaded: {filename} ({{len(file_bytes)}} bytes)")
"""

            response = code_interpreter.invoke("executeCode", {
                "code": write_code,
                "language": "python",
                "clearContext": False
            })

            # Check for errors
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = result.get("structuredContent", {}).get("stderr", "Unknown error")
                    logger.error(f"Failed to upload image {filename}: {error_msg[:200]}")
                    continue

            uploaded.append(filename)
            logger.info(f"✅ Uploaded image to Code Interpreter: {filename}")

        except Exception as e:
            logger.error(f"Failed to upload image {filename}: {e}")
            continue

    return uploaded


@tool(context=True)
def create_excel_spreadsheet(
    python_code: str,
    spreadsheet_name: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Create a new Excel spreadsheet using openpyxl code.

    This tool executes openpyxl code to create a spreadsheet from scratch.
    Perfect for generating structured data with sheets, tables, charts, and formatting.

    Available libraries: openpyxl, pandas, matplotlib, numpy

    Use this tool when:
    - User asks to create/generate a new Excel spreadsheet
    - User wants a spreadsheet with specific data structure
    - User needs charts/pivot tables in the initial spreadsheet

    Args:
        python_code: Python code using openpyxl to build the spreadsheet.
                    The workbook is initialized as: wb = Workbook()
                    The active sheet is: ws = wb.active
                    After your code, it's automatically saved.

                    DO NOT include Workbook() initialization or wb.save() calls.

                    Uploaded images are automatically available in Code Interpreter.
                    Use os.listdir() to discover available image files.

                    Common Patterns:

                    Basic Data Entry:
                    ```python
# Set sheet title
ws.title = 'Sales Data'

# Add headers with formatting
ws['A1'] = 'Product'
ws['B1'] = 'Quantity'
ws['C1'] = 'Price'
from openpyxl.styles import Font, PatternFill
ws['A1'].font = Font(bold=True, size=12)
ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Add data rows
data = [
    ['Widget A', 100, 25.50],
    ['Widget B', 150, 30.00],
    ['Widget C', 200, 20.00]
]
for row in data:
    ws.append(row)
                    ```

                    With Table:
                    ```python
from openpyxl.worksheet.table import Table, TableStyleInfo

# Add data first
ws['A1'] = 'Product'
ws['B1'] = 'Sales'
ws.append(['Product A', 1000])
ws.append(['Product B', 1500])

# Create table
tab = Table(displayName='SalesTable', ref='A1:B3')
style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)
                    ```

                    With Chart:
                    ```python
from openpyxl.chart import BarChart, Reference

# Add data
ws['A1'] = 'Month'
ws['B1'] = 'Sales'
for i, (month, sales) in enumerate([('Jan', 100), ('Feb', 120), ('Mar', 150)], 2):
    ws[f'A{i}'] = month
    ws[f'B{i}'] = sales

# Create chart
chart = BarChart()
chart.title = 'Monthly Sales'
chart.x_axis.title = 'Month'
chart.y_axis.title = 'Sales'

data = Reference(ws, min_col=2, min_row=1, max_row=4)
categories = Reference(ws, min_col=1, min_row=2, max_row=4)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

ws.add_chart(chart, 'D2')
                    ```

                    With Uploaded Image:
                    ```python
from openpyxl.drawing.image import Image
import os

# Discover available images
available_images = [f for f in os.listdir() if f.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))]

if available_images:
    # Use the first image (or select specific one by filename matching)
    image_file = available_images[0]
    img = Image(image_file)
    ws.add_image(img, 'E1')
else:
    # No images found - add note in cell
    ws['E1'] = '[No images available in workspace]'
                    ```

                    With Multiple Sheets:
                    ```python
# Create additional sheets
ws2 = wb.create_sheet('Summary')
ws3 = wb.create_sheet('Details')

# Add data to each sheet
ws.title = 'Overview'
ws['A1'] = 'Main Data'

ws2['A1'] = 'Summary'
ws3['A1'] = 'Detailed Analysis'
                    ```

        spreadsheet_name: Spreadsheet name WITHOUT extension (.xlsx is added automatically)
                         Use ONLY letters, numbers, hyphens (no underscores or spaces)
                         Examples: "sales-report", "Q4-data", "inventory-2024"

    Returns:
        Success message with file details and workspace list

    Note:
        - Spreadsheet is saved to workspace for future editing
        - Uploaded images are automatically available in Code Interpreter
        - Keep code focused on structure; use modify_excel_spreadsheet for refinements
    """
    try:
        logger.info("=== create_excel_spreadsheet called ===")
        logger.info(f"Spreadsheet name: {spreadsheet_name}")

        # Validate spreadsheet name (without extension)
        is_valid, error_msg = _validate_spreadsheet_name(spreadsheet_name)
        if not is_valid:
            return {
                "content": [{
                    "text": f"❌ **Invalid spreadsheet name**: {spreadsheet_name}\n\n{error_msg}\n\n**Examples of valid names:**\n- sales-report\n- Q4-data\n- inventory-2024"
                }],
                "status": "error"
            }

        # Add .xlsx extension
        spreadsheet_filename = f"{spreadsheet_name}.xlsx"
        logger.info(f"Full filename: {spreadsheet_filename}")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelDocumentManager(user_id, session_id)

        # Get Code Interpreter
        code_interpreter_id = _get_code_interpreter_id()
        if not code_interpreter_id:
            return {
                "content": [{
                    "text": "❌ **Code Interpreter not configured**\n\nCODE_INTERPRETER_ID not found in environment or Parameter Store."
                }],
                "status": "error"
            }

        region = os.getenv('AWS_REGION', 'us-west-2')
        code_interpreter = CodeInterpreter(region)
        code_interpreter.start(identifier=code_interpreter_id)

        try:
            # Load all workspace images from S3 to Code Interpreter
            loaded_images = _load_workspace_images_to_ci(code_interpreter, user_id, session_id)
            if loaded_images:
                logger.info(f"Loaded {len(loaded_images)} image(s) from workspace: {loaded_images}")

            # Get Code Interpreter path for file
            ci_path = doc_manager.get_ci_path(spreadsheet_filename)

            # Build spreadsheet creation code
            creation_code = f"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# Create new workbook
wb = Workbook()
ws = wb.active

# Execute user's creation code
{python_code}

# Save workbook
wb.save('{ci_path}')
print(f"Spreadsheet created: {ci_path}")
"""

            # Execute creation
            response = code_interpreter.invoke("executeCode", {
                "code": creation_code,
                "language": "python",
                "clearContext": False
            })

            # Check for errors
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = result.get("structuredContent", {}).get("stderr", "Unknown error")
                    logger.error(f"Creation failed: {error_msg[:500]}")
                    code_interpreter.stop()
                    return {
                        "content": [{
                            "text": f"❌ **Failed to create spreadsheet**\n\n```\n{error_msg[:1000]}\n```\n\n💡 Check your openpyxl code for syntax errors or incorrect API usage."
                        }],
                        "status": "error"
                    }

            logger.info("Spreadsheet creation completed")

            # Download from Code Interpreter
            file_bytes = doc_manager.download_from_code_interpreter(code_interpreter, spreadsheet_filename)

            # Save to S3 for persistence
            s3_info = doc_manager.save_to_s3(
                spreadsheet_filename,
                file_bytes,
                metadata={'source': 'python_code_creation'}
            )

            # Get current workspace list
            workspace_docs = doc_manager.list_s3_documents()
            workspace_summary = doc_manager.format_file_list(workspace_docs)

            message = f"""✅ **Spreadsheet created successfully**

**File**: {spreadsheet_filename} ({s3_info['size_kb']})

{workspace_summary}"""

            # Return success message
            return {
                "content": [{"text": message}],
                "status": "success",
                "metadata": {
                    "filename": spreadsheet_filename,
                    "tool_type": "excel_spreadsheet",
                    "user_id": user_id,
                    "session_id": session_id
                }
            }

        finally:
            code_interpreter.stop()

    except Exception as e:
        logger.error(f"create_excel_spreadsheet failed: {e}")
        return {
            "content": [{
                "text": f"❌ **Failed to create spreadsheet**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def modify_excel_spreadsheet(
    source_name: str,
    output_name: str,
    python_code: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Modify existing Excel spreadsheet using openpyxl code and save with a new name.

    This tool loads a spreadsheet from workspace, executes openpyxl code to modify it,
    and saves it with a new filename to preserve the original.

    Available libraries: openpyxl, pandas, matplotlib, numpy

    Use this tool when:
    - User wants to edit/modify/update an existing spreadsheet
    - User asks to add sheets, data, charts, or images to a spreadsheet
    - User wants to refine or change parts of a spreadsheet

    IMPORTANT Safety Rules:
    - Always use different output_name than source_name (e.g., "report" → "report-v2")
    - Always check sheet exists before accessing
    - Use try-except for operations that might fail

    Args:
        source_name: Spreadsheet name to load (WITHOUT extension, must exist in workspace)
                    Example: "sales-report", "Q4-data"
        output_name: New spreadsheet name (WITHOUT extension, must be different from source)
                    Use ONLY letters, numbers, hyphens (no underscores or spaces)
                    Example: "sales-report-v2", "Q4-data-final"
        python_code: Python code using openpyxl library to modify spreadsheet.
                    The workbook is loaded as: wb = load_workbook('<filename>')
                    After modifications, it's automatically saved.

                    DO NOT include load_workbook() or wb.save() calls.

                    IMPORTANT: Uploaded images are automatically available in Code Interpreter.

                    Common Patterns:

                    Add New Sheet with Data:
                    ```python
# Create new sheet
ws_new = wb.create_sheet('Q1 Summary')

# Add data
ws_new['A1'] = 'Summary Data'
ws_new['A1'].font = Font(bold=True, size=14)
ws_new.append(['Item', 'Value'])
ws_new.append(['Total', 10000])
                    ```

                    Modify Existing Data:
                    ```python
# Access existing sheet
ws = wb['Sales Data']

# Update specific cells
ws['B2'] = 150  # Update value
ws['C2'].font = Font(color='FF0000')  # Change color
                    ```

                    Add Chart to Existing Sheet:
                    ```python
from openpyxl.chart import BarChart, Reference

# Access sheet
ws = wb.active

# Create chart from existing data
chart = BarChart()
chart.title = 'Sales Analysis'
data = Reference(ws, min_col=2, min_row=1, max_row=10)
categories = Reference(ws, min_col=1, min_row=2, max_row=10)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

ws.add_chart(chart, 'E2')
                    ```

                    Insert Uploaded Image:
                    ```python
from openpyxl.drawing.image import Image

# Access sheet
ws = wb['Dashboard']

# Images from workspace are automatically available
import os
images = [f for f in os.listdir() if f.endswith(('.png', '.jpg', '.jpeg'))]
if images:
    img = Image(images[0])
    ws.add_image(img, 'F5')
                    ```

    Returns:
        Success message with file details and workspace list

    Note:
        - Uploaded images are automatically available in Code Interpreter
        - Document automatically synced to S3
    """
    try:
        logger.info("=== modify_excel_spreadsheet called ===")
        logger.info(f"Source: {source_name}, Output: {output_name}")

        # Validate output name format
        is_valid, error_msg = _validate_spreadsheet_name(output_name)
        if not is_valid:
            return {
                "content": [{
                    "text": f"❌ **Invalid output name**: {output_name}\n\n{error_msg}\n\n**Examples of valid names:**\n- sales-report-v2\n- Q4-data-final\n- report-revised"
                }],
                "status": "error"
            }

        # Ensure source and output are different
        if source_name == output_name:
            return {
                "content": [{
                    "text": f"❌ **Invalid name**\n\nOutput name must be different from source name to preserve the original.\n\nSource: {source_name}\nOutput: {output_name}\n\n💡 Try: \"{source_name}-v2\""
                }],
                "status": "error"
            }

        # Add .xlsx extensions
        source_filename = f"{source_name}.xlsx"
        output_filename = f"{output_name}.xlsx"
        logger.info(f"Full filenames: {source_filename} → {output_filename}")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelDocumentManager(user_id, session_id)

        # Get Code Interpreter
        code_interpreter_id = _get_code_interpreter_id()
        if not code_interpreter_id:
            return {
                "content": [{
                    "text": "❌ **Code Interpreter not configured**\n\nCODE_INTERPRETER_ID not found in environment or Parameter Store."
                }],
                "status": "error"
            }

        region = os.getenv('AWS_REGION', 'us-west-2')
        code_interpreter = CodeInterpreter(region)
        code_interpreter.start(identifier=code_interpreter_id)

        try:
            # Load all workspace images from S3 to Code Interpreter
            loaded_images = _load_workspace_images_to_ci(code_interpreter, user_id, session_id)
            if loaded_images:
                logger.info(f"Loaded {len(loaded_images)} image(s) from workspace: {loaded_images}")

            # Ensure source file is in Code Interpreter (load from S3 if needed)
            source_ci_path = doc_manager.ensure_file_in_ci(code_interpreter, source_filename)

            # Generate output path
            output_ci_path = doc_manager.get_ci_path(output_filename)

            # Build modification code
            modification_code = f"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load source spreadsheet
wb = load_workbook('{source_ci_path}')

# Execute user's modification code
{python_code}

# Save to output file
wb.save('{output_ci_path}')
print(f"Spreadsheet modified and saved: {output_ci_path}")
"""

            # Execute modification
            response = code_interpreter.invoke("executeCode", {
                "code": modification_code,
                "language": "python",
                "clearContext": False
            })

            # Check for errors
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = result.get("structuredContent", {}).get("stderr", "Unknown error")
                    logger.error(f"Modification failed: {error_msg[:500]}")
                    code_interpreter.stop()
                    return {
                        "content": [{
                            "text": f"❌ **Modification failed**\n\n```\n{error_msg[:1000]}\n```\n\n💡 Check your openpyxl code for syntax errors or incorrect API usage."
                        }],
                        "status": "error"
                    }

            logger.info("Spreadsheet modification completed")

            # Download modified spreadsheet from Code Interpreter
            file_bytes = doc_manager.download_from_code_interpreter(code_interpreter, output_filename)

            # Save to S3 with output filename
            s3_info = doc_manager.save_to_s3(
                output_filename,
                file_bytes,
                metadata={
                    'source': 'modification',
                    'source_filename': source_filename,
                    'modified_at': 'timestamp'
                }
            )

            # Get current workspace list
            workspace_docs = doc_manager.list_s3_documents()
            workspace_summary = doc_manager.format_file_list(workspace_docs)

            # Build success message
            message = f"""✅ **Spreadsheet modified successfully**

**Source**: {source_filename}
**Saved as**: {output_filename} ({s3_info['size_kb']})

{workspace_summary}"""

            # Return success message with metadata for download button
            return {
                "content": [{"text": message}],
                "status": "success",
                "metadata": {
                    "filename": output_filename,
                    "tool_type": "excel_spreadsheet",
                    "user_id": user_id,
                    "session_id": session_id
                }
            }

        finally:
            code_interpreter.stop()

    except FileNotFoundError as e:
        logger.error(f"Spreadsheet not found: {e}")
        return {
            "content": [{
                "text": f"❌ **Spreadsheet not found**: {source_filename}"
            }],
            "status": "error"
        }
    except Exception as e:
        logger.error(f"modify_excel_spreadsheet failed: {e}")
        return {
            "content": [{
                "text": f"❌ **Failed to modify spreadsheet**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def list_my_excel_spreadsheets(
    tool_context: ToolContext
) -> Dict[str, Any]:
    """List all Excel spreadsheets in workspace.

    Shows all .xlsx files in workspace with size and metadata.

    Use this tool when:
    - User asks "what Excel files do I have?"
    - User says "show my spreadsheets", "list files"
    - Before modifying: verify spreadsheet exists
    - User wants to see workspace contents

    No arguments needed.

    Returns:
        - Formatted list of all Excel spreadsheets
        - Each entry shows: filename, size, last modified date
        - Total file count
        - Metadata for frontend download buttons

    Example Usage:
        Scenario 1 - Check available files:
            User: "What Excel spreadsheets do I have?"
            AI: list_my_excel_spreadsheets()
            → Shows: sales.xlsx, inventory.xlsx, report.xlsx

        Scenario 2 - Before modifying:
            User: "Edit my sales data"
            AI: [Unclear which file]
            AI: list_my_excel_spreadsheets()
            AI: "I found these spreadsheets: ... Which one should I modify?"

    Example Output:
        📁 Workspace (3 spreadsheets):
          - sales-report.xlsx (52.3 KB) - Modified: 2025-01-15
          - inventory.xlsx (41.8 KB) - Modified: 2025-01-14
          - Q4-analysis.xlsx (89.2 KB) - Modified: 2025-01-13

    Note:
        - Shows files from workspace
        - Empty workspace shows helpful message
        - Frontend renders download buttons automatically
    """
    try:
        logger.info("=== list_my_excel_spreadsheets called ===")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelDocumentManager(user_id, session_id)

        # List documents from S3
        documents = doc_manager.list_s3_documents()

        # Format list
        workspace_summary = doc_manager.format_file_list(documents)

        if documents:
            message = workspace_summary
        else:
            message = workspace_summary

        # Prepare metadata for frontend (download buttons)
        metadata = {
            "documents": [
                {
                    "filename": doc['filename'],
                    "s3_key": doc['s3_key'],
                    "size_kb": doc['size_kb'],
                    "last_modified": doc['last_modified']
                } for doc in documents
            ]
        }

        return {
            "content": [{"text": message}],
            "status": "success",
            "metadata": metadata
        }

    except Exception as e:
        logger.error(f"list_my_excel_spreadsheets failed: {e}")
        return {
            "content": [{
                "text": f"❌ **Failed to list spreadsheets**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def read_excel_spreadsheet(
    spreadsheet_name: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Read and retrieve a specific Excel spreadsheet.

    This tool loads a spreadsheet from workspace and returns it as downloadable bytes.
    The spreadsheet content is accessible to you (the agent) for analysis and answering questions.

    Use this tool when:
    - User asks about spreadsheet contents: "What's in sales.xlsx?", "Summarize this data"
    - User wants to analyze the spreadsheet: "How many sheets?", "What's the total?"
    - User explicitly requests download: "Send me [filename]", "I need [spreadsheet]"
    - You need to verify spreadsheet contents before modification

    IMPORTANT:
    - For creating new spreadsheets: use create_excel_spreadsheet
    - For modifying spreadsheets: use modify_excel_spreadsheet

    Args:
        spreadsheet_name: Spreadsheet name WITHOUT extension (.xlsx is added automatically)
                         Must exist in workspace.
                         Example: "sales-report", "inventory", "Q4-data"

    Returns:
        - Spreadsheet metadata (filename, size, S3 location)
        - Special metadata format for frontend download
        - Frontend automatically shows download button

    Example Usage:
        # Download request
        User: "Send me the sales report"
        AI: read_excel_spreadsheet("sales-report")

        # After creation
        User: "Create report and send it"
        AI: create_excel_spreadsheet(...)
        AI: read_excel_spreadsheet("sales-report")

    Note:
        - File must exist in workspace
        - Frontend handles download automatically
    """
    try:
        logger.info("=== read_excel_spreadsheet called ===")
        logger.info(f"Spreadsheet name: {spreadsheet_name}")

        # Add .xlsx extension
        spreadsheet_filename = f"{spreadsheet_name}.xlsx"
        logger.info(f"Full filename: {spreadsheet_filename}")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelDocumentManager(user_id, session_id)

        # Load from S3
        file_bytes = doc_manager.load_from_s3(spreadsheet_filename)

        # Get file info
        documents = doc_manager.list_s3_documents()
        doc_info = next((d for d in documents if d['filename'] == spreadsheet_filename), None)

        if not doc_info:
            raise FileNotFoundError(f"Spreadsheet not found: {spreadsheet_filename}")

        message = f"""✅ **Spreadsheet ready for download**

**File**: {spreadsheet_filename} ({doc_info['size_kb']})
**Last Modified**: {doc_info['last_modified'].split('T')[0]}"""

        # Sanitize spreadsheet name for Bedrock API (remove extension, handle legacy files)
        sanitized_name = _sanitize_spreadsheet_name_for_bedrock(spreadsheet_filename)

        # Return with downloadable bytes
        return {
            "content": [
                {"text": message},
                {
                    "document": {
                        "format": "xlsx",
                        "name": sanitized_name,
                        "source": {
                            "bytes": file_bytes
                        }
                    }
                }
            ],
            "status": "success",
            "metadata": {
                "filename": spreadsheet_filename,
                "s3_key": doc_manager.get_s3_key(spreadsheet_filename),
                "size_kb": doc_info['size_kb'],
                "last_modified": doc_info['last_modified'],
                "tool_type": "excel_spreadsheet",
                "user_id": user_id,
                "session_id": session_id
            }
        }

    except FileNotFoundError as e:
        logger.error(f"Spreadsheet not found: {e}")
        return {
            "content": [{
                "text": f"❌ **Spreadsheet not found**: {spreadsheet_filename}"
            }],
            "status": "error"
        }
    except Exception as e:
        logger.error(f"read_excel_spreadsheet failed: {e}")
        return {
            "content": [{
                "text": f"❌ **Failed to read spreadsheet**\n\n{str(e)}"
            }],
            "status": "error"
        }
